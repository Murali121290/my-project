from __future__ import annotations
import csv
import io
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from manuscript_core.exporters import BODY_FONT, HEADER_FILL, HEADER_FONT, _autofit

MATCH_COLORS = [
    "FFFF00", "90EE90", "FFB6C1", "FFA500", "E6E6FA",
    "FFDAB9", "87CEEB", "98FB98", "F08080", "00FFFF",
]

FILE_BANNER_FILL = PatternFill(start_color="1A2742", end_color="1A2742", fill_type="solid")
FILE_BANNER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def load_file_data(path: str) -> tuple[list[str], list[list[Any]]]:
    ext = Path(path).suffix.lower()
    if ext == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return [], []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        data = [list(r) for r in rows[1:]]
        return headers, data
    elif ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return [], []
        headers = rows[0]
        data = [list(r) for r in rows[1:]]
        return headers, data
    else:
        raise ValueError(f"Unsupported file type: {ext}. Only .xlsx and .csv are supported.")


def build_column_map(all_headers: list[list[str]]) -> list[str]:
    seen: dict[str, int] = {}
    ordered: list[str] = []
    for headers in all_headers:
        for h in headers:
            key = h.strip()
            if key not in seen:
                seen[key] = len(ordered)
                ordered.append(key)
    return ordered


def _normalize(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def build_merged_comparison_workbook(
    file_paths: list[str],
    filenames: list[str],
) -> io.BytesIO:
    all_headers: list[list[str]] = []
    all_data: list[list[list[Any]]] = []

    for path in file_paths:
        headers, data = load_file_data(path)
        all_headers.append(headers)
        all_data.append(data)

    master_cols = build_column_map(all_headers)
    max_col = len(master_cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "Merged & Compared"

    current_row = 1

    # Single master header row at the top
    for ci, col_name in enumerate(master_cols, start=1):
        c = ws.cell(row=current_row, column=ci, value=col_name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
    current_row += 1

    # row_map: list of (ws_row, file_index, normalized_fingerprint_tuple)
    row_map: list[tuple[int, int, tuple]] = []

    for fi, (_, headers, data) in enumerate(zip(filenames, all_headers, all_data)):
        # Build column-name → position map for this file
        col_index = {h.strip(): i for i, h in enumerate(headers)}

        # Data rows
        for row_vals in data:
            norm_parts: list[str] = []
            for col_name in master_cols:
                src_idx = col_index.get(col_name.strip())
                val = row_vals[src_idx] if src_idx is not None and src_idx < len(row_vals) else None
                c = ws.cell(row=current_row, column=master_cols.index(col_name) + 1, value=val)
                c.font = BODY_FONT
                c.alignment = WRAP
                norm_parts.append(_normalize(val))
            fp = tuple(norm_parts)
            row_map.append((current_row, fi, fp))
            current_row += 1

    _apply_row_highlights(ws, row_map, max_col)
    summary_data = _build_summary_data(row_map, filenames, master_cols)
    _write_summary_sheet(wb, summary_data)
    _autofit(ws, {i: 40 for i in range(1, max_col + 1)})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _apply_row_highlights(ws, row_map, max_col):
    fingerprint_groups: dict[tuple, list[tuple[int, int]]] = defaultdict(list)

    for (ws_row, fi, fp) in row_map:
        if all(v == "" for v in fp):
            continue
        fingerprint_groups[fp].append((ws_row, fi))

    color_index = 0
    for fp, occurrences in fingerprint_groups.items():
        distinct_files = {fi for (_, fi) in occurrences}
        if len(distinct_files) < 2:
            continue

        hex_color = MATCH_COLORS[color_index % len(MATCH_COLORS)]
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        for (ws_row, _) in occurrences:
            for col in range(1, max_col + 1):
                ws.cell(row=ws_row, column=col).fill = fill
        color_index += 1


def _build_summary_data(
    row_map: list[tuple[int, int, tuple]],
    filenames: list[str],
    master_cols: list[str],
) -> list[dict]:
    fingerprint_groups: dict[tuple, list[tuple[int, int]]] = defaultdict(list)
    for (ws_row, fi, fp) in row_map:
        if all(v == "" for v in fp):
            continue
        fingerprint_groups[fp].append((ws_row, fi))

    results = []
    color_index = 0
    for fp, occurrences in fingerprint_groups.items():
        distinct_files = {fi for (_, fi) in occurrences}
        if len(distinct_files) < 2:
            continue
        hex_color = MATCH_COLORS[color_index % len(MATCH_COLORS)]
        file_names_str = ", ".join(filenames[fi] for fi in sorted(distinct_files))
        preview = fp[0] if fp else ""
        results.append({
            "color": hex_color,
            "files": file_names_str,
            "count": len(occurrences),
            "preview": preview,
        })
        color_index += 1
    return results


def _write_summary_sheet(wb: Workbook, summary_data: list[dict]) -> None:
    ws = wb.create_sheet(title="Duplicate Summary")
    headers = ["Color", "Matched Files", "Row Count", "Preview (first column value)"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")

    for ri, entry in enumerate(summary_data, start=2):
        swatch_fill = PatternFill(
            start_color=entry["color"], end_color=entry["color"], fill_type="solid"
        )
        label_cell = ws.cell(row=ri, column=1, value=entry["color"])
        label_cell.fill = swatch_fill
        label_cell.font = Font(name="Arial", size=10)
        ws.cell(row=ri, column=2, value=entry["files"]).font = BODY_FONT
        ws.cell(row=ri, column=3, value=entry["count"]).font = BODY_FONT
        ws.cell(row=ri, column=4, value=entry["preview"]).font = BODY_FONT

    _autofit(ws, {1: 15, 2: 50, 3: 12, 4: 40})
