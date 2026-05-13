"""
Manuscript Consistency Checker Blueprint for PPH.
Integrates the manuscript_consistency project as a Flask blueprint with two workflows:
1. Analysis / Editorial Stylesheet Prep (PM role) → dashboard
2. Technical Editing (COPYEDIT role) → per-occurrence review
"""
from __future__ import annotations

import io
import json
import uuid
import zipfile
import tempfile
import shutil
from datetime import datetime
from pathlib import Path
from functools import wraps

from flask import (
    Blueprint, request, session, render_template, redirect, url_for,
    flash, send_file, jsonify, Response, g
)
from werkzeug.utils import secure_filename

from manuscript_core.analyzer import analyze_manuscript
from manuscript_core.exporters import build_combined_excel, build_csv, build_excel, build_ia_excel
from manuscript_core.fixer import apply_fixes_to_docx, apply_fixes_targeted, build_fixes_from_selection
from manuscript_core.tasks import analyze_job, fix_job

# Try to import RQ/Redis for async support
try:
    from redis import Redis
    from rq import Queue
    from rq.job import Job
    REDIS_AVAILABLE = True
    redis_conn = Redis(decode_responses=True)
    task_queue = Queue(connection=redis_conn)
except Exception:
    REDIS_AVAILABLE = False
    task_queue = None

# ============ Paths ============
BASE_DIR = Path(__file__).parent
MANU_UPLOAD_DIR  = BASE_DIR / 'S4C-Processed-Documents' / 'manuscript_uploads'
MANU_RESULTS_DIR = BASE_DIR / 'S4C-Processed-Documents' / 'manuscript_results'
MANU_OUTPUTS_DIR = BASE_DIR / 'S4C-Processed-Documents' / 'manuscript_outputs'
MANU_IA_DIR      = BASE_DIR / 'S4C-Processed-Documents' / 'manuscript_ia_mappings'

for _d in [MANU_UPLOAD_DIR, MANU_RESULTS_DIR, MANU_OUTPUTS_DIR, MANU_IA_DIR]:
    _d.mkdir(parents=True, exist_ok=True)

# Templates data file
TEMPLATES_FILE = BASE_DIR / 'manuscript_core' / 'data' / 'templates.json'

ALLOWED_EXT = {".docx"}
MAX_CONTENT_LENGTH = 64 * 1024 * 1024  # 64 MB

# ============ Blueprint ============
manuscript_bp = Blueprint(
    'manuscript',
    __name__,
    template_folder='templates/manuscript',
)

ALLOWED_ROLES = {'COPYEDIT', 'COPYEDITPM', 'PM', 'ADMIN'}

# ============ Auth Decorator ============
def manuscript_auth_required(f):
    """Auth guard for manuscript routes."""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            flash("Please log in to continue.")
            return redirect(url_for('login'))
        role = (session.get('role') or '').upper()
        if not session.get('is_admin') and role not in ALLOWED_ROLES:
            flash("You don't have permission to access this page.", "error")
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return wrapped

# ============ Helpers ============
def _allowed(filename: str) -> bool:
    """Check if file extension is allowed."""
    return Path(filename).suffix.lower() in ALLOWED_EXT

def _load_results(job_id: str) -> dict | None:
    """Load analysis results by job_id."""
    # First check for flat file (legacy format)
    flat_path = MANU_RESULTS_DIR / f"{job_id}.json"
    if flat_path.exists():
        try:
            data = json.loads(flat_path.read_text(encoding="utf-8"))
            if data.get("job_id") == job_id:
                return data
        except (json.JSONDecodeError, IOError):
            pass

    # Search in subfolders
    try:
        for item in MANU_RESULTS_DIR.iterdir():
            if not item.is_dir():
                continue
            result_path = item / "results.json"
            if result_path.is_file():
                try:
                    data = json.loads(result_path.read_text(encoding="utf-8"))
                    if data.get("job_id") == job_id:
                        return data
                except (json.JSONDecodeError, IOError):
                    continue
    except (OSError, IOError):
        pass

    return None


def _sanitize(name: str) -> str:
    """Sanitize a name for use as a filename component."""
    return "".join(c for c in name if c.isalnum() or c in "-_").strip() or name


def _ensure_ia_mapping_from_selection(
    client_name: str,
    project_name: str,
    ia_mapping_path: Path
) -> bool:
    """If ia_mapping.py doesn't exist, try to create it from a saved selection (JSON or DB)."""
    if ia_mapping_path.exists():
        return True

    def _rows_to_ia_file(selected_rows: list[dict]) -> bool:
        ia_rows = [
            (row.get("element", ""), row.get("subtype", ""), row.get("pattern", ""), None)
            for row in selected_rows
            if row.get("element") and row.get("pattern")
        ]
        if not ia_rows:
            return False
        ia_mapping_path.parent.mkdir(parents=True, exist_ok=True)
        lines = ["IA_TEMPLATE_ROWS = ["]
        for r in ia_rows:
            lines.append(f"    {r},")
        lines.append("]\n")
        ia_mapping_path.write_text("\n".join(lines), encoding="utf-8")
        return True

    safe_project = _sanitize(project_name)
    safe_client = _sanitize(client_name)

    # Strategy 1: Try naming convention candidates
    candidates = [
        MANU_IA_DIR / f"{safe_project}_{safe_client}_rules.json",
        MANU_IA_DIR / f"{safe_client}_{safe_project}_rules.json",
        MANU_IA_DIR / f"{project_name}_{client_name}_rules.json",
        MANU_IA_DIR / f"{client_name}_{project_name}_rules.json",
    ]
    for candidate in candidates:
        if candidate.exists():
            try:
                sel_data = json.loads(candidate.read_text(encoding="utf-8"))
                if _rows_to_ia_file(sel_data.get("selected_ia_rows", [])):
                    return True
            except Exception:
                continue

    # Strategy 2: Scan ALL JSON selection files in the directory
    try:
        if MANU_IA_DIR.exists():
            for json_file in MANU_IA_DIR.glob("*_*_rules.json"):
                try:
                    sel_data = json.loads(json_file.read_text(encoding="utf-8"))
                    sp = sel_data.get("project_name", "")
                    sc = sel_data.get("client_name", "")
                    if sp in (project_name, client_name, safe_project, safe_client) or \
                       sc in (project_name, client_name, safe_project, safe_client) or \
                       not (project_name or client_name):
                        if _rows_to_ia_file(sel_data.get("selected_ia_rows", [])):
                            return True
                except Exception:
                    continue
    except Exception:
        pass

    # Strategy 3: Desperate — pick most recent JSON file regardless of name
    try:
        if MANU_IA_DIR.exists():
            json_files = sorted(MANU_IA_DIR.glob("*_*_rules.json"), key=lambda p: p.stat().st_mtime, reverse=True)
            if json_files:
                sel_data = json.loads(json_files[0].read_text(encoding="utf-8"))
                if _rows_to_ia_file(sel_data.get("selected_ia_rows", [])):
                    return True
    except Exception:
        pass

    return False


# ============ Routes: Upload Pages ============

@manuscript_bp.route('/analysis', methods=['GET'])
@manuscript_auth_required
def analysis_upload():
    """Analysis / Editorial Stylesheet Prep upload page (COPYEDITPM only)."""
    role = (session.get('role') or '').upper()
    if role != 'COPYEDITPM' and not session.get('is_admin'):
        flash("This page is for Analysis only. Redirecting to Technical Editing.", "info")
        return redirect(url_for('manuscript.technical_upload'))

    return render_template('analysis_upload.html', current_role=g.user.get('role') if g.user else None)

@manuscript_bp.route('/technical-edit', methods=['GET'])
@manuscript_auth_required
def technical_upload():
    """Technical Editing upload page (COPYEDIT / COPYEDITPM only)."""
    role = (session.get('role') or '').upper()
    if role not in ['COPYEDIT', 'COPYEDITPM'] and not session.get('is_admin'):
        flash("This page is for Technical Editing only. Redirecting to Analysis.", "info")
        return redirect(url_for('manuscript.analysis_upload'))

    # Load templates list
    templates = []
    if TEMPLATES_FILE.exists():
        try:
            data = json.loads(TEMPLATES_FILE.read_text(encoding="utf-8"))
            templates = data.get("templates", [])
        except (json.JSONDecodeError, IOError):
            pass

    return render_template('technical_upload.html', templates=templates, current_role=g.user.get('role') if g.user else None)

# ============ Routes: Analysis ============

@manuscript_bp.route('/analyze', methods=['POST'])
@manuscript_auth_required
def analyze():
    """Analyze uploaded manuscript chapters."""
    client_name = request.form.get('client_name') or ""
    project_name = request.form.get('project_name') or request.form.get('template')
    files = request.files.getlist("word_files[]")

    if not files or all(not f.filename for f in files):
        return jsonify({"error": "No files were uploaded."}), 400

    # Create folder naming: ClientName_ProjectName_YYYYMMDD_HHMMSS
    job_id = uuid.uuid4().hex[:10]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    client_project = f"{client_name}_{project_name}".strip("_")
    if not client_project:
        client_project = "manuscript"
    folder_name = f"{client_project}_{timestamp}"

    job_dir = MANU_UPLOAD_DIR / folder_name
    job_dir.mkdir(parents=True, exist_ok=True)

    saved_chapters = []
    # Create IA mapping path: manuscript_ia_mappings/ClientName_ProjectName/ia_mapping.py
    # Use a more robust path construction
    client_name_safe = _sanitize(client_name) if client_name else "unknown"
    project_name_safe = _sanitize(project_name) if project_name else "unknown"
    
    # Construct ia_folder with fallback logic
    if client_name and project_name:
        ia_folder = f"{client_name}_{project_name}"
    elif client_name:
        ia_folder = client_name
    elif project_name:
        ia_folder = project_name
    else:
        ia_folder = "default"
    
    ia_mapping_path = MANU_IA_DIR / ia_folder / "ia_mapping.py"

    # Ensure template file exists before analysis (so analyzer.py can use the correct rows)
    if project_name and client_name and (session.get('role') or '').upper() != 'PM':
        _ensure_ia_mapping_from_selection(client_name, project_name, ia_mapping_path)
    elif not ia_mapping_path.exists():
        # Create minimal default ia_mapping if file doesn't exist
        try:
            ia_mapping_path.parent.mkdir(parents=True, exist_ok=True)
            ia_mapping_path.write_text("IA_TEMPLATE_ROWS = []\n", encoding="utf-8")
        except Exception:
            pass  # Non-critical, will use DEFAULT_ROWS in analyzer

    for idx, f in enumerate(sorted(files, key=lambda x: x.filename or ""), start=1):
        if not f.filename or not _allowed(f.filename):
            continue
        safe_name = secure_filename(f.filename)
        dest = job_dir / safe_name
        f.save(dest)
        saved_chapters.append(
            {
                "index": idx,
                "filename": safe_name,
                "path": str(dest),
                "client_name": client_name,
                "project_name": project_name,
                "role": session.get("role"),
                "ia_mapping_path": str(ia_mapping_path)
            }
        )

    if not saved_chapters:
        return jsonify({"error": "No valid .docx files were uploaded."}), 400

    # Queue analysis job (or run synchronously if Redis unavailable)
    if REDIS_AVAILABLE and task_queue:
        try:
            rq_job = task_queue.enqueue(analyze_job, job_id, saved_chapters, job_timeout='10m')
            return jsonify({"job_id": job_id, "task_id": rq_job.id, "status": "queued"})
        except Exception as e:
            pass  # Fall through to sync

    # Fallback: synchronous analysis
    try:
        findings = analyze_manuscript(saved_chapters)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Analysis failed: {type(e).__name__}: {e}"}), 500

    # Add job_id to results
    findings["job_id"] = job_id

    # Save results in organized structure
    results_dir = MANU_RESULTS_DIR / folder_name
    results_dir.mkdir(parents=True, exist_ok=True)
    out_path = results_dir / "results.json"
    out_path.write_text(json.dumps(findings, ensure_ascii=False), encoding="utf-8")

    # Auto-save outputs
    job_out_dir = MANU_OUTPUTS_DIR / folder_name
    job_out_dir.mkdir(parents=True, exist_ok=True)

    # 1. Excel report
    try:
        xlsx_bytes = build_excel(findings, job_id)
        (job_out_dir / f"{client_project}_manuscript_consistency_report.xlsx").write_bytes(xlsx_bytes)
    except Exception as _exc:
        pass

    # 2. Standalone dashboard HTML
    try:
        role = session.get("role")
        ia_mapping_path_str = str(MANU_IA_DIR / ia_folder / "ia_mapping.py")
        html = render_template(
            "dashboard_standalone.html",
            data=findings,
            job_id=job_id,
            standalone=True,
            role=role,
            ia_mapping_path=ia_mapping_path_str,
            current_role=role
        )
        (job_out_dir / f"{client_project}_dashboard.html").write_text(html, encoding="utf-8")
    except Exception as _exc:
        pass

    # Update templates list
    if project_name:
        try:
            if TEMPLATES_FILE.exists():
                data = json.loads(TEMPLATES_FILE.read_text(encoding="utf-8"))
            else:
                data = {"templates": []}

            if project_name not in data["templates"]:
                data["templates"].append(project_name)

            TEMPLATES_FILE.write_text(json.dumps(data, indent=4, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    return jsonify({"job_id": job_id})

# ============ Routes: Dashboard / Results ============

@manuscript_bp.route('/dashboard/<job_id>', methods=['GET'])
@manuscript_auth_required
def dashboard(job_id: str):
    """Analysis results dashboard."""
    data = _load_results(job_id)
    if data is None:
        flash("Job not found.", "error")
        return redirect(url_for('manuscript.analysis_upload'))

    from manuscript_core.ia_mapping import RULE_ID_TO_IA
    if "ia_report" in data and "rule_id_to_ia" not in data["ia_report"]:
        data["ia_report"]["rule_id_to_ia"] = RULE_ID_TO_IA

    chapters = data.get("chapters", [])
    ia_mapping_path = chapters[0].get("ia_mapping_path", "") if chapters else ""
    project_name = chapters[0].get("project_name", "") if chapters else ""
    client_name = chapters[0].get("client_name", "") if chapters else ""

    return render_template(
        "manuscript_dashboard.html",
        data=data,
        job_id=job_id,
        session_id=job_id,
        project_name=project_name,
        client_name=client_name,
        standalone=False,
        role=g.user.get("role") if g.user else None,
        current_role=g.user.get("role") if g.user else None,
        ia_mapping_path=ia_mapping_path
    )

@manuscript_bp.route('/review/<job_id>', methods=['GET'])
@manuscript_auth_required
def editor_review(job_id: str):
    """Technical Editor per-occurrence review page."""
    data = _load_results(job_id)
    if data is None:
        flash("Job not found.", "error")
        return redirect(url_for('manuscript.technical_upload'))

    chapters = data.get("chapters", [])
    if not chapters:
        flash("No chapters found.", "error")
        return redirect(url_for('manuscript.technical_upload'))

    # Get ia_mapping_path from first chapter
    ia_mapping_path_str = chapters[0].get("ia_mapping_path")
    if not ia_mapping_path_str:
        flash("Template path not found.", "error")
        return redirect(url_for('manuscript.technical_upload'))

    ia_mapping_path = Path(ia_mapping_path_str)

    # If ia_mapping.py doesn't exist, try to recover from saved selection
    if not ia_mapping_path.exists():
        client_name = chapters[0].get("client_name", "")
        project_name = chapters[0].get("project_name", "")
        # Try multiple approaches to recover the ia_mapping file
        if not _ensure_ia_mapping_from_selection(client_name, project_name, ia_mapping_path):
            # As last resort, ensure parent directory exists and create a minimal template
            try:
                ia_mapping_path.parent.mkdir(parents=True, exist_ok=True)
                ia_mapping_path.write_text(
                    "IA_TEMPLATE_ROWS = []\n",
                    encoding="utf-8"
                )
            except Exception:
                flash("Template file not found and could not be created. Ask the PM to save a rule selection first.", "error")
                return redirect(url_for('manuscript.technical_upload'))

    try:
        from manuscript_core.ia_mapping import RULE_ID_TO_IA
        import re
        import json as _json

        with open(ia_mapping_path, encoding="utf-8") as f:
            ia_data = {}
            exec(f.read(), ia_data)
        IA_TEMPLATE_ROWS = ia_data.get("IA_TEMPLATE_ROWS", [])

        # Ensure findings have correct encoding
        all_findings_raw = data.get("findings", [])
        # Re-encode to ensure proper UTF-8
        try:
            all_findings = _json.loads(_json.dumps(all_findings_raw, ensure_ascii=False, default=str))
        except:
            all_findings = all_findings_raw

        # Determine which rules match the template
        preferred_patterns = {}
        for element, _, pattern, _ in IA_TEMPLATE_ROWS:
            if element and pattern:
                preferred_patterns.setdefault(element, set()).add(pattern)

        target_rule_ids = set()
        for rule_id, (mapped_element, _, mapped_pattern) in RULE_ID_TO_IA.items():
            if mapped_element in preferred_patterns:
                # If the element was selected AT ALL, we add this rule ID to ensure all findings for this element are shown
                target_rule_ids.add(rule_id)

        # Allow dynamic spelling rules if any spelling preference was selected
        spelling_active = "American spellings" in preferred_patterns or "British spellings" in preferred_patterns

        # Show only findings that match the target_rule_ids based on the user's selection
        # EXCEPT if the user has no preferred_patterns (meaning they didn't select anything in the UI, show everything)
        findings = []

        for f in all_findings:
            rule_id = f.get("rule_id", "")
            
            # Dynamic matching for spelling
            if rule_id.startswith("spelling_") and spelling_active:
                findings.append(f)
                continue
                
            if rule_id in target_rule_ids or not preferred_patterns:
                rule_id = f.get("rule_id")
                surface = f.get("surface", "")

                # Dynamic replacement for ranges
                if rule_id in ("range_to", "range_endash", "range_hyphen"):
                    prefs = preferred_patterns.get("Ranges", set())
                    if prefs:
                        pref = list(prefs)[0]
                        nums = re.findall(r'\d+', surface)
                        if len(nums) >= 2:
                            if "to" in pref:
                                f["replacement"] = f"{nums[0]} to {nums[1]}"
                            elif "en dash" in pref:
                                f["replacement"] = f"{nums[0]}–{nums[1]}"
                            elif "hyphen" in pref:
                                f["replacement"] = f"{nums[0]}–{nums[1]}"

                # Dynamic replacement for thousand separators
                elif rule_id in ("thous_sep_missing", "thous_sep_comma", "thous_sep_space", "thous_sep_nbsp"):
                    prefs = preferred_patterns.get("Thousand separator (use/non-use)", set())
                    if prefs:
                        pref = list(prefs)[0]
                        clean_num = re.sub(r'[,\s\u00a0]', '', surface)
                        try:
                            if "comma" in pref:
                                f["replacement"] = f"{int(clean_num):,}"
                            elif "no comma" in pref:
                                f["replacement"] = clean_num
                        except ValueError:
                            pass

                findings.append(f)

        # Sort: TE points first, remaining after
        te_findings = [f for f in findings if f.get("category") == "te_point"]
        other_findings = [f for f in findings if f.get("category") != "te_point"]
        sorted_findings = te_findings + other_findings
        te_count = len(te_findings)

        return render_template(
            "editor_review.html",
            findings=sorted_findings,
            te_count=te_count,
            job_id=job_id,
            chapters=chapters,
            current_role=g.user.get("role") if g.user else None
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Error loading template: {str(e)}", "error")
        return redirect(url_for('manuscript.technical_upload'))

# ============ Routes: Fix Operations ============

@manuscript_bp.route('/fix/<job_id>', methods=['POST'])
@manuscript_auth_required
def fix_document(job_id: str):
    """Apply pattern-level fixes."""
    data = _load_results(job_id)
    if data is None:
        return jsonify({"error": "Job not found"}), 404

    # Find job directory
    job_dir = None
    chapters = data.get("chapters", [])
    if chapters:
        first_chapter_filename = chapters[0]["filename"]
        for potential_dir in MANU_UPLOAD_DIR.glob("*/"):
            if (potential_dir / first_chapter_filename).exists():
                job_dir = potential_dir
                break

    if not job_dir or not job_dir.exists():
        return jsonify({"error": "Files not found"}), 404

    req_data = request.json or {}
    selected_patterns = req_data.get("selected_patterns", [])

    # Queue fix job (or run synchronously)
    if REDIS_AVAILABLE and task_queue:
        try:
            rq_job = task_queue.enqueue(fix_job, job_id, selected_patterns, data['findings'], data['chapters'], job_timeout='15m')
            return jsonify({"task_id": rq_job.id, "status": "queued"})
        except Exception as e:
            pass  # Fall through to sync

    # Fallback: synchronous fix
    from manuscript_core.figure_table_highlighter import FigureTableHighlighter
    
    fixes = build_fixes_from_selection(selected_patterns, data)
    
    # Check if we need to apply highlighting
    highlight_elements = set()
    for pat in selected_patterns:
        elem = pat.get("element")
        if elem in ("Figure", "Table", "Box", "Exhibit", "Appendix", "Case Study"):
            highlight_elements.add(elem)
            
    if not fixes and not highlight_elements:
        return jsonify({"error": "No auto-fix rules or highlight rules apply to the selected patterns."}), 400

    temp_dir = Path(tempfile.mkdtemp())
    zip_path = temp_dir / f"Fixed_Manuscript_{job_id}.zip"

    try:
        highlighter = FigureTableHighlighter() if highlight_elements else None
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for chapter in data["chapters"]:
                orig_file = job_dir / chapter["filename"]
                if orig_file.exists():
                    fixed_name = f"FIXED_{chapter['filename']}"
                    fixed_file = temp_dir / fixed_name
                    
                    # 1. Apply auto-fixes (track changes)
                    if fixes:
                        apply_fixes_to_docx(orig_file, fixed_file, fixes)
                    else:
                        # Copy original if no fixes
                        import shutil
                        shutil.copy2(orig_file, fixed_file)
                        
                    # 2. Apply highlighting (modifies fixed_file in-place)
                    if highlighter and highlight_elements:
                        highlighter.apply_highlighting_to_docx(
                            str(fixed_file), 
                            str(fixed_file), 
                            list(highlight_elements)
                        )
                        
                    zip_ref.write(fixed_file, fixed_name)

        return send_file(
            path_or_file=zip_path,
            mimetype="application/zip",
            as_attachment=True,
            download_name=zip_path.name
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

@manuscript_bp.route('/fix-by-id/<job_id>', methods=['POST'])
@manuscript_auth_required
def fix_by_id(job_id: str):
    """Apply targeted per-occurrence fixes."""
    data = _load_results(job_id)
    if data is None:
        return jsonify({"error": "Job not found"}), 404

    req_data = request.json or {}
    selected_findings = req_data.get("selected_findings", [])
    highlight_findings = req_data.get("highlight_findings", [])

    if not selected_findings and not highlight_findings:
        return jsonify({"error": "No findings selected."}), 400

    # Find job directory
    job_dir = None
    chapters = data.get("chapters", [])
    if chapters:
        first_chapter_filename = chapters[0]["filename"]
        for potential_dir in MANU_UPLOAD_DIR.glob("*/"):
            if (potential_dir / first_chapter_filename).exists():
                job_dir = potential_dir
                break

    if not job_dir or not job_dir.exists():
        return jsonify({"error": "Source files not found"}), 404

    # Build per-chapter highlight_texts from highlight_findings
    import re as _re
    from manuscript_core.fixer import apply_te_highlights_to_docx

    def _build_hl_texts(raw_findings):
        seen = set()
        result = []
        for f in raw_findings:
            pat_str = f.get("search_pattern")
            if not pat_str:
                surface = f.get("surface", "")
                if not surface:
                    continue
                pat_str = r'\b' + _re.escape(surface) + r'\b'
            key = (pat_str, f.get("region", "body"))
            if key in seen:
                continue
            seen.add(key)
            result.append({
                "pattern":       _re.compile(pat_str, _re.IGNORECASE),
                "region":        f.get("region", "body"),
                "source_filter": f.get("source", "body"),
                "rule_id":       f.get("rule_id", ""),
                "surface":       f.get("surface", ""),
            })
        return result

    temp_dir = Path(tempfile.mkdtemp())
    zip_path = temp_dir / f"Fixed_Manuscript_{job_id}.zip"

    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for chapter in data["chapters"]:
                ch_idx = chapter["index"]
                ch_fixes = [sf for sf in selected_findings if sf.get("chapter_index") == ch_idx]
                ch_hl_raw = [hf for hf in highlight_findings if hf.get("chapter_index") == ch_idx]
                ch_hl_texts = _build_hl_texts(ch_hl_raw)
                orig_file = job_dir / chapter["filename"]
                if not orig_file.exists():
                    continue
                fixed_name = f"FIXED_{chapter['filename']}"
                fixed_file = temp_dir / fixed_name
                if ch_fixes:
                    apply_fixes_targeted(orig_file, fixed_file, ch_fixes)
                else:
                    shutil.copy2(orig_file, fixed_file)
                if ch_hl_texts:
                    apply_te_highlights_to_docx(str(fixed_file), str(fixed_file), ch_hl_texts)
                zip_ref.write(fixed_file, fixed_name)

        return send_file(
            path_or_file=zip_path,
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"Fixed_Manuscript_{job_id}.zip",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

# ============ Routes: IA Mapping ============

@manuscript_bp.route('/save-ia', methods=['POST'])
@manuscript_auth_required
def save_ia():
    """Save IA mapping configuration."""
    data = request.json
    rows = data.get("data", [])
    file_path_str = data.get("path")

    if not file_path_str:
        return jsonify({"status": "error", "message": "Path is required"}), 400

    if not isinstance(rows, list):
        return jsonify({"status": "error", "message": "Invalid data format"}), 400

    file_path = Path(file_path_str)

    # Ensure directory exists
    try:
        file_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Cannot create directory: {str(e)}"}), 500

    # Version old file if exists
    version_created = None
    if file_path.exists():
        version = 1
        while True:
            versioned_path = file_path.parent / f"{file_path.stem}_v{version}{file_path.suffix}"
            if not versioned_path.exists():
                try:
                    file_path.rename(versioned_path)
                    version_created = version
                    break
                except Exception as e:
                    return jsonify({"status": "error", "message": f"Cannot version old file: {str(e)}"}), 500
            version += 1

    # Clean data
    IA_TEMPLATE_ROWS_CLEAN = [
        (r[0], r[1], r[2], r[3] if r[3] not in ("", None) else None)
        for r in rows
        if r[0] and r[1] and r[2]
    ]

    if not IA_TEMPLATE_ROWS_CLEAN:
        return jsonify({
            "status": "error",
            "message": "No valid patterns selected."
        }), 400

    # Save file
    try:
        with file_path.open("w", encoding="utf-8") as f:
            f.write("IA_TEMPLATE_ROWS = [\n")
            for i, row in enumerate(IA_TEMPLATE_ROWS_CLEAN):
                comma = "," if i < len(IA_TEMPLATE_ROWS_CLEAN) - 1 else ""
                f.write(f"    {row}{comma}\n")
            f.write("]\n")
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Unable to save: {str(e)}"
        }), 500

    return jsonify({
        "status": "success",
        "file": file_path.name,
        "versioned": version_created
    })

# ============ Routes: Downloads ============

@manuscript_bp.route('/download/<job_id>/report.xlsx', methods=['GET'])
def download_excel(job_id: str):
    """Download Excel report."""
    data = _load_results(job_id)
    if data is None:
        return jsonify({"error": "not found"}), 404
    xlsx_bytes = build_excel(data, job_id)
    return send_file(
        path_or_file=io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"manuscript_consistency_{job_id}.xlsx",
    )

@manuscript_bp.route('/download/<job_id>/ia_report.xlsx', methods=['GET'])
def download_ia_excel(job_id: str):
    """Download IA-format Excel report."""
    data = _load_results(job_id)
    if data is None:
        return jsonify({"error": "not found"}), 404
    xlsx_bytes = build_ia_excel(data, job_id)
    return send_file(
        path_or_file=io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"ia_report_{job_id}.xlsx",
    )

@manuscript_bp.route('/download/<job_id>/combined_report.xlsx', methods=['GET'])
def download_combined_excel(job_id: str):
    """Download combined Excel report (consistency + IA report in one workbook)."""
    data = _load_results(job_id)
    if data is None:
        return jsonify({"error": "not found"}), 404
    xlsx_bytes = build_combined_excel(data, job_id)
    return send_file(
        path_or_file=io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"manuscript_report_{job_id}.xlsx",
    )

@manuscript_bp.route('/download/<job_id>/findings.csv', methods=['GET'])
def download_csv(job_id: str):
    """Download CSV findings."""
    data = _load_results(job_id)
    if data is None:
        return jsonify({"error": "not found"}), 404
    csv_bytes = build_csv(data)
    return send_file(
        path_or_file=io.BytesIO(csv_bytes),
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=f"manuscript_findings_{job_id}.csv",
    )

@manuscript_bp.route('/download/<job_id>/sheet.html', methods=['GET'])
@manuscript_bp.route('/sheet/<job_id>', methods=['GET'])
def download_dashboard(job_id: str):
    """Download standalone dashboard HTML."""
    data = _load_results(job_id)
    if data is None:
        return jsonify({"error": "not found"}), 404
    html = render_template(
        "dashboard_standalone.html",
        data=data,
        job_id=job_id,
        standalone=True,
        role=session.get("role"),
        current_role=g.user.get("role") if g.user else None,
        ia_mapping_path=data["chapters"][0].get("ia_mapping_path", ""),
    )
    return Response(
        html,
        mimetype="text/html; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="dashboard_{job_id}.html"',
        },
    )

# ============ Routes: API & Status ============

@manuscript_bp.route('/api/results/<job_id>', methods=['GET'])
def api_results(job_id: str):
    """Get raw JSON results."""
    data = _load_results(job_id)
    if data is None:
        return jsonify({"error": "not found"}), 404
    return jsonify(data)

@manuscript_bp.route('/job-status/<task_id>', methods=['GET'])
def job_status(task_id: str):
    """Check status of a queued task."""
    if not REDIS_AVAILABLE or not task_queue:
        return jsonify({"error": "Job queue not available"}), 503

    try:
        job = Job.fetch(task_id, connection=redis_conn)
        if job.is_finished:
            return jsonify({"status": "completed", "result": job.result})
        elif job.is_failed:
            return jsonify({"status": "failed", "error": str(job.exc_info)})
        else:
            return jsonify({"status": "processing"})
    except Exception as e:
        return jsonify({"status": "unknown", "error": str(e)}), 404


# ============ Routes: Discovery UI ============

@manuscript_bp.route('/discovery', methods=['GET'])
@manuscript_auth_required
def discovery():
    """Discovery UI page for rule selection."""
    return render_template("discovery.html")


@manuscript_bp.route('/discovery/<session_id>/ia-rows', methods=['GET'])
@manuscript_auth_required
def discovery_ia_rows(session_id: str):
    """List all IA template rows with detected counts."""
    data = _load_results(session_id)
    if data is None:
        return jsonify({"error": "Session not found"}), 404

    ia_report = data.get("ia_report", {})
    ia_rows = ia_report.get("rows", [])

    ia_rows_with_counts = []
    elements_set = set()
    for row in ia_rows:
        element = row.get("element", "")
        elements_set.add(element)
        ia_rows_with_counts.append({
            "element": element,
            "subtype": row.get("type", ""),
            "pattern": row.get("pattern", ""),
            "example": row.get("example") or "",
            "detected_count": row.get("total", 0),
            "found": row.get("total", 0) > 0,
        })

    return jsonify({
        "ia_rows": ia_rows_with_counts,
        "elements": sorted(elements_set),
    })


@manuscript_bp.route('/discovery/<session_id>/create-selection', methods=['POST'])
@manuscript_auth_required
def create_selection(session_id: str):
    """Create and save a rule selection to a JSON file."""
    try:
        data = request.get_json()
        selection_name = data.get("selection_name", "").strip()
        description = data.get("description", "").strip()
        selected_ia_rows = data.get("selected_ia_rows", [])
        custom_grouping = data.get("custom_grouping", {})
        project_name = data.get("project_name", "").strip()
        client_name = data.get("client_name", "").strip()

        if not selection_name:
            return jsonify({"error": "selection_name required"}), 400

        selection_id = str(uuid.uuid4())

        # Save to JSON file
        try:
            ia_mappings_dir = Path(BASE_DIR) / "S4C-Processed-Documents" / "manuscript_ia_mappings"
            ia_mappings_dir.mkdir(parents=True, exist_ok=True)

            safe_project = _sanitize(project_name) or "Project"
            safe_client = _sanitize(client_name) or "Client"
            filename = f"{safe_project}_{safe_client}_rules.json"
            filepath = ia_mappings_dir / filename

            selection_data = {
                "id": selection_id,
                "selection_name": selection_name,
                "description": description,
                "project_name": project_name,
                "client_name": client_name,
                "session_id": session_id,
                "selected_ia_rows": selected_ia_rows,
                "custom_grouping": custom_grouping,
                "created_at": str(datetime.now()),
                "created_by": session.get("username", "unknown"),
                "active": False,
            }

            filepath.write_text(json.dumps(selection_data, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"✓ Selection saved to file: {filepath}")
        except Exception as file_error:
            print(f"Warning: Could not save selection file: {file_error}")

        # Also create ia_mapping.py so technical editing can find the template
        try:
            if selected_ia_rows and client_name and project_name:
                ia_folder = f"{client_name}_{project_name}".strip("_") or project_name
                ia_mapping_path = ia_mappings_dir / ia_folder / "ia_mapping.py"
                ia_mapping_path.parent.mkdir(parents=True, exist_ok=True)
                ia_rows = [
                    (row.get("element", ""), row.get("subtype", ""), row.get("pattern", ""), None)
                    for row in selected_ia_rows
                    if row.get("element") and row.get("pattern")
                ]
                if ia_rows:
                    lines = ["IA_TEMPLATE_ROWS = ["]
                    for r in ia_rows:
                        lines.append(f"    {r},")
                    lines.append("]\n")
                    ia_mapping_path.write_text("\n".join(lines), encoding="utf-8")
                    print(f"✓ Template also saved to: {ia_mapping_path}")
        except Exception as tmpl_error:
            print(f"Warning: Could not save ia_mapping.py: {tmpl_error}")

        return jsonify({
            "selection_id": selection_id,
            "status": "saved",
            "file_saved": True,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@manuscript_bp.route('/discovery/<session_id>/ia-report', methods=['GET'])
@manuscript_auth_required
def discovery_ia_report(session_id: str):
    """Generate filtered IA report as Excel download."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    selection_id = request.args.get("selection_id")
    if not selection_id:
        return jsonify({"error": "selection_id required"}), 400

    _, sel_data = _find_selection_json(selection_id)
    if not sel_data:
        return jsonify({"error": "Selection not found"}), 404

    data = _load_results(session_id)
    if data is None:
        return jsonify({"error": "Session not found"}), 404

    ia_report = data.get("ia_report", {})
    all_ia_rows = ia_report.get("rows", [])
    chapter_indices = ia_report.get("chapter_indices", [])
    chapter_names = ia_report.get("chapter_names", {})

    selected_patterns = sel_data.get("selected_ia_rows", [])
    selected_set = {(r.get("element"), r.get("subtype"), r.get("pattern")) for r in selected_patterns}

    filtered_rows = [
        row for row in all_ia_rows
        if (row.get("element"), row.get("type"), row.get("pattern")) in selected_set
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "IA Report"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="00408E", end_color="00408E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)

    headers = ["Element", "Type", "Pattern", "Example"] + \
              [chapter_names.get(str(i), f"Ch{i}") for i in chapter_indices] + ["Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r_idx, row in enumerate(filtered_rows, 2):
        ws.cell(row=r_idx, column=1, value=row.get("element", ""))
        ws.cell(row=r_idx, column=2, value=row.get("type", ""))
        ws.cell(row=r_idx, column=3, value=row.get("pattern", ""))
        ws.cell(row=r_idx, column=4, value=row.get("example", ""))
        by_ch = row.get("by_chapter", {})
        total = 0
        for i_idx, i in enumerate(chapter_indices, 5):
            count = by_ch.get(str(i), 0)
            ws.cell(row=r_idx, column=i_idx, value=count)
            total += count
        ws.cell(row=r_idx, column=len(headers), value=total)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"ia_report_selection_{selection_id}.xlsx",
    )


@manuscript_bp.route('/rule-selections', methods=['GET'])
@manuscript_auth_required
def rule_selections_list():
    """List saved rule selections page."""
    return render_template("rule_selections.html")


@manuscript_bp.route('/rule-selections/api', methods=['GET'])
@manuscript_auth_required
def rule_selections_api():
    """JSON API listing all saved selections (from JSON files)."""
    session_id = request.args.get("session_id")
    try:
        selections = []
        if MANU_IA_DIR.exists():
            files = sorted(MANU_IA_DIR.glob("*_rules.json"), key=lambda p: p.stat().st_mtime, reverse=True)
            for f in files[:50]:
                try:
                    s = json.loads(f.read_text(encoding="utf-8"))
                    if session_id and s.get("session_id") != session_id:
                        continue
                    selections.append({
                        "id": s.get("id", ""),
                        "selection_name": s.get("selection_name", ""),
                        "description": s.get("description", ""),
                        "project_name": s.get("project_name", ""),
                        "client_name": s.get("client_name", ""),
                        "session_id": s.get("session_id", ""),
                        "num_rules": len(s.get("selected_ia_rows", [])),
                        "created_at": s.get("created_at", ""),
                        "created_by": s.get("created_by", ""),
                        "active": s.get("active", False),
                    })
                except Exception:
                    continue
        return jsonify({"selections": selections})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _find_selection_json(selection_id: str) -> tuple:
    """Return (path, data) for the JSON file whose 'id' matches, or (None, None)."""
    if not MANU_IA_DIR.exists():
        return None, None
    for f in MANU_IA_DIR.glob("*_rules.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            if data.get("id") == selection_id:
                return f, data
        except Exception:
            continue
    return None, None


@manuscript_bp.route('/rule-selections/<selection_id>/activate', methods=['POST'])
@manuscript_auth_required
def activate_selection(selection_id: str):
    """Activate a specific selection and deactivate others for the same project/client."""
    try:
        target_path, target_data = _find_selection_json(selection_id)
        if not target_data:
            return jsonify({"error": "Selection not found"}), 404

        project_name = target_data.get("project_name", "")
        client_name = target_data.get("client_name", "")

        # Deactivate all JSON files with same project/client, then activate the target
        for f in MANU_IA_DIR.glob("*_rules.json"):
            try:
                d = json.loads(f.read_text(encoding="utf-8"))
                if d.get("project_name") == project_name and d.get("client_name") == client_name:
                    d["active"] = (d.get("id") == selection_id)
                    f.write_text(json.dumps(d, indent=2, ensure_ascii=False), encoding="utf-8")
            except Exception:
                continue

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@manuscript_bp.route('/rule-selections/<selection_id>', methods=['DELETE'])
@manuscript_auth_required
def delete_selection(selection_id: str):
    """Delete a rule selection JSON file."""
    try:
        target_path, _ = _find_selection_json(selection_id)
        if not target_path:
            return jsonify({"error": "Selection not found"}), 404
        target_path.unlink()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@manuscript_bp.route('/rule-selections/<selection_id>', methods=['PUT'])
@manuscript_auth_required
def update_selection(selection_id: str):
    """Update a rule selection's metadata in its JSON file."""
    try:
        req_data = request.json
        name = req_data.get("selection_name")
        if not name:
            return jsonify({"error": "Selection name required"}), 400

        target_path, target_data = _find_selection_json(selection_id)
        if not target_data:
            return jsonify({"error": "Selection not found"}), 404

        target_data["selection_name"] = name
        target_data["description"] = req_data.get("description", target_data.get("description", ""))
        target_data["project_name"] = req_data.get("project_name", target_data.get("project_name", ""))
        target_data["client_name"] = req_data.get("client_name", target_data.get("client_name", ""))
        target_path.write_text(json.dumps(target_data, indent=2, ensure_ascii=False), encoding="utf-8")

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
